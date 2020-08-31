import os
import win32com.client
from openpyxl import Workbook
import openpyxl


def excel_handler(path, date_name):

    new_folder = 'Отчеты'
    new_path = os.path.join(path, new_folder)
    if not os.path.exists(new_path):
        os.makedirs(new_path)
    os.chdir(new_path)
    # objs = obj_base
    objs = []

    format_file = '.xlsx'
    filename = date_name + format_file
    path = os.path.join(new_path, filename)

    wb = Workbook()
    wb.create_sheet(date_name, 0)
    wb.save(filename)

    excel = win32com.client.Dispatch("Excel.Application")
    work_b1 = excel.Workbooks.Open(path)

    sheet = work_b1.Worksheets(1)
    sheet.Cells(2, 1).ColumnWidth = 6
    sheet.Cells(2, 2).ColumnWidth = 4
    sheet.Cells(2, 3).ColumnWidth = 10
    sheet.Cells(2, 4).ColumnWidth = 15
    sheet.Cells(2, 5).ColumnWidth = 10
    sheet.Cells(2, 6).ColumnWidth = 10
    sheet.Cells(2, 7).ColumnWidth = 10
    sheet.Cells(2, 8).ColumnWidth = 10
    sheet.Cells(2, 9).ColumnWidth = 10
    Selection = sheet.Range(sheet.Cells(2, 6), sheet.Cells(4, 9))
    Selection.Merge()

    Selection = sheet.Range(sheet.Cells(5, 6), sheet.Cells(5, 7))
    Selection.Merge()

    Selection = sheet.Range(sheet.Cells(5, 8), sheet.Cells(5, 9))
    Selection.Merge()

    for i in range(1, 6):
        Selection = sheet.Range(sheet.Cells(2, i), sheet.Cells(6, i))
        Selection.Merge()

    sheet.Cells(2, 1).Value = 'Дата'
    sheet.Cells(2, 2).Value = '№ п/п'
    sheet.Cells(2, 3).Value = 'Заказчик'
    sheet.Cells(2, 4).Value = 'Адрес места сбора и накопления ТКО и КГО'
    sheet.Cells(2, 5).Value = 'Время загрузки, час-минуты'
    sheet.Cells(2, 6).Value = 'Количество ТКО с мест сбора и накопления, осуществляемого контейнерным способом'

    sheet.Cells(5, 6).Value = 'объем, м3'
    sheet.Cells(5, 8).Value = 'количество контейнеров, шт'
    sheet.Cells(6, 6).Value = 'План'
    sheet.Cells(6, 7).Value = 'Факт'
    sheet.Cells(6, 8).Value = 'План'
    sheet.Cells(6, 9).Value = 'Факт'

    Selection = sheet.Range("A2:I6")
    Selection.WrapText = True
    Selection.Font.Size = 11
    Selection.Borders.Weight = 2
    Selection.HorizontalAlignment = -4108
    Selection.VerticalAlignment = -4108

    work_b1.Save()
    work_b1.Close()
    excel.Quit()


if __name__ == '__main__':
    path = os.getcwd()
    excel_handler(path, date_name='20.08.2020')

    # sheet.Cells(1, 8).ColumnWidth = 12
    # sheet.Cells(1, 9).ColumnWidth = 12
    # sheet.Cells(1, 10).ColumnWidth = 14
    # sheet.Cells(1, 11).ColumnWidth = 18
    # sheet.Cells(1, 12).ColumnWidth = 12
    # sheet.Cells(1, 13).ColumnWidth = 12
    #
    # sheet.Cells(1, 14).ColumnWidth = 20
    # sheet.Cells(1, 15).ColumnWidth = 8
    # sheet.Cells(1, 16).ColumnWidth = 10
    # sheet.Cells(1, 17).ColumnWidth = 19
    # sheet.Cells(1, 18).ColumnWidth = 12
    # sheet.Cells(1, 19).ColumnWidth = 12
    # sheet.Cells(1, 20).ColumnWidth = 12
    # sheet.Cells(1, 21).ColumnWidth = 10
    # sheet.Cells(1, 22).ColumnWidth = 13
    #
    # sheet.Cells(1, 23).ColumnWidth = 13
    # sheet.Cells(1, 24).ColumnWidth = 22
    # sheet.Cells(1, 25).ColumnWidth = 7
    #
    # sheet.Cells(1, 26).ColumnWidth = 13
    # sheet.Cells(1, 27).ColumnWidth = 22
    # sheet.Cells(1, 28).ColumnWidth = 7
    #
    # sheet.Cells(1, 29).ColumnWidth = 16
    #
    # sheet.Cells(1, 8).Value = 'Сводный отчет/Моточасы'
    #
    # Selection = sheet.Range("H1:M1")
    # Selection.Borders.Weight = 2
    # Selection.Interior.Color = 5296274
    # Selection.HorizontalAlignment = -4108
    # Selection.Font.Size = 11
    # Selection.Merge()
    #
    # sheet.Cells(1, 14).Value = 'Отчет по рейсам'
    #
    # Selection = sheet.Range("N1:V1")
    # Selection.Borders.Weight = 2
    # Selection.Interior.Color = 16763904
    # Selection.HorizontalAlignment = -4108
    # Selection.Font.Size = 11
    # Selection.Merge()
    #
    # sheet.Cells(1, 23).Value = 'Простой больше 15 мин.'
    #
    # Selection = sheet.Range("W1:Y1")
    # Selection.Borders.Weight = 2
    # Selection.Interior.Color = 6724095
    # Selection.HorizontalAlignment = -4108
    # Selection.Font.Size = 11
    # Selection.Merge()
    #
    # sheet.Cells(1, 26).Value = 'Простой больше часа'
    #
    # Selection = sheet.Range("Z1:AB1")
    # Selection.Borders.Weight = 2
    # Selection.Interior.Color = 6724095
    # Selection.HorizontalAlignment = -4108
    # Selection.Font.Size = 11
    # Selection.Merge()
    #
    # sheet.Cells(1, 29).Value = 'Всего простоев'
    # Selection = sheet.Range("AC1:AC2")
    # Selection.Borders.Weight = 2
    # Selection.HorizontalAlignment = -4108
    # Selection.Font.Size = 11
    # Selection.Merge()
    #
    # Selection = sheet.Range("A2:AB2")
    # Selection.Borders.Weight = 2
    # Selection.Font.Size = 10
    # Selection.Font.Bold = True
    # Selection.RowHeight = 13
    # Selection.Interior.Color = 15921906
    # Selection.HorizontalAlignment = -4108



    # sheet.Cells(2, 14).Value = 'Маршрут'
    # sheet.Cells(2, 15).Value = 'Кол-во'
    # sheet.Cells(2, 16).Value = 'Пробег'
    # sheet.Cells(2, 17).Value = 'Длительность поездки'
    # sheet.Cells(2, 18).Value = 'Ср. скорость'
    # sheet.Cells(2, 19).Value = 'Макс. скорость'
    # sheet.Cells(2, 20).Value = 'Потрачено'
    # sheet.Cells(2, 21).Value = 'Длительность стоянок'
    # sheet.Cells(2, 22).Value = 'Ср. расход'
    #
    # sheet.Cells(2, 23).Value = 'Длительность'
    # sheet.Cells(2, 24).Value = 'Положение'
    # sheet.Cells(2, 25).Value = 'Кол-во'
    #
    # sheet.Cells(2, 26).Value = 'Длительность'
    # sheet.Cells(2, 27).Value = 'Положение'
    # sheet.Cells(2, 28).Value = 'Кол-во'

        # work_b1.Save()
        # work_b1.Close()
        # excel.Quit()
        #
        # wb = openpyxl.load_workbook(filename)
        # ws = wb.worksheets(sheet_count)
        # ws.auto_filter.ref = 'A2:Y2'
        # wb.save(filename)
        #
        # excel = win32com.client.Dispatch("Excel.Application")
        # work_b1 = excel.Workbooks.Open(path)
        # sheet = work_b1.Worksheets(sheet_count)

    # i = stroka
    #
    # for obj in objs:
    #
    #     sheet.Cells(i, 1).Value = date
    #     sheet.Cells(i, 2).Value = obj.mark
    #     sheet.Cells(i, 3).Value = obj.gos_number
    #
    #     if smena == 1:
    #         sheet.Cells(i, 4).Value = 'Смена 1'
    #     else:
    #         sheet.Cells(i, 4).Value = 'Смена 2'
    #
    #     sheet.Cells(i, 6).Value = obj.fuel_up_sum
    #     sheet.Cells(i, 7).Value = obj.milliage
    #
    #     sheet.Cells(i, 8).Value = obj.moto_hours
    #     sheet.Cells(i, 9).Value = obj.move_time
    #     sheet.Cells(i, 10).Value = obj.moto_hh
    #     sheet.Cells(i, 11).Value = obj.fuel_dut
    #     sheet.Cells(i, 12).Value = obj.fuel_start
    #     sheet.Cells(i, 13).Value = obj.fuel_end
    #
    #     sheet.Cells(i, 23).Value = obj.prost_15_duration
    #     sheet.Cells(i, 24).Value = obj.prost_15_place
    #     sheet.Cells(i, 25).Value = obj.prost_15_count
    #
    #     sheet.Cells(i, 26).Value = obj.prost_1h_duration
    #     sheet.Cells(i, 27).Value = obj.prost_1h_place
    #     sheet.Cells(i, 28).Value = obj.prost_1h_count
    #
    #     sheet.Cells(i, 29).Value = obj.prost_full_duration
    #
    #     flight_k = 0
    #     start = i
    #
    #     if obj.fligt_total > 0:
    #         for _ in range(obj.fligt_total):
    #             sheet.Cells(i, 1).Value = date
    #             sheet.Cells(i, 2).Value = obj.mark
    #             sheet.Cells(i, 3).Value = obj.gos_number
    #
    #             if smena == 1:
    #                 sheet.Cells(i, 4).Value = 'Смена 1'
    #             else:
    #                 sheet.Cells(i, 4).Value = 'Смена 2'
    #
    #             sheet.Cells(i, 14).Value = obj.flight_target[flight_k]
    #             sheet.Cells(i, 15).Value = obj.flight_count[flight_k]
    #             sheet.Cells(i, 16).Value = obj.flight_mileage[flight_k]
    #             sheet.Cells(i, 17).Value = obj.flight_duration[flight_k]
    #             sheet.Cells(i, 18).Value = obj.flight_avg_speed[flight_k]
    #             sheet.Cells(i, 19).Value = obj.flight_max_speed[flight_k]
    #             sheet.Cells(i, 20).Value = obj.flight_spent[flight_k]
    #             sheet.Cells(i, 21).Value = obj.flight_duration_stop[flight_k]
    #             sheet.Cells(i, 22).Value = obj.flight_avg_spent[flight_k]
    #
    #             flight_k += 1
    #             if obj.fligt_total > 1:
    #                 i += 1
    #
    #     if obj.fligt_total > 1:
    #         i -= 1
    #
    #     if flight_k > 1:
    #         for j in range(9):
    #             point_a = sheet.Cells(start, j + 5)
    #             point_b = sheet.Cells(i, j + 5)
    #             Selection = sheet.Range(point_a, point_b)
    #             Selection.Merge()
    #
    #         for j in range(23, 30):
    #             point_a = sheet.Cells(start, j)
    #             point_b = sheet.Cells(i, j)
    #             Selection = sheet.Range(point_a, point_b)
    #             Selection.Merge()
    #     i += 1
    #
    # i -= 1
    # A1 = sheet.Cells(3, 1)
    # A2 = sheet.Cells(i, 29)
    # Selection = sheet.Range(A1, A2)
    # Selection.Font.Size = 10
    # Selection.RowHeight = 13
    # Selection.HorizontalAlignment = -4152
    # Selection.Borders.Weight = 2
    #
    # k = str(i)
    #
    # Selection = sheet.Range('K3:M' + k)
    # Selection.NumberFormat = '0,#0" "л'
    #
    # Selection = sheet.Range('G3:G' + k)
    # Selection.NumberFormat = '0,#0" "к"м"'
    #
    # Selection = sheet.Range('F3:F' + k)
    # Selection.NumberFormat = '0,#0" "л'
    #
    # Selection = sheet.Range('AC3:AC' + k)
    # Selection.NumberFormat = '[ч]" часов "м" минут"'
    #
    # Selection = sheet.Range('P3:P' + k)
    # Selection.NumberFormat = '0,#0" "к"м"'
    #
    # Selection = sheet.Range('T3:T' + k)
    # Selection.NumberFormat = '0,#0" "л'

